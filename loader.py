# loader.py
import os, random, glob, re
import streamlit as st
from openpyxl import load_workbook
from helpers import parse_duration

class Loader:
    def __init__(self, project_root, valence_condition):
        from config import AFFECT_EXCEL, STIMULI_EXCEL
        self.project_root = project_root
        self.affect_excel = AFFECT_EXCEL
        self.stimuli_excel = STIMULI_EXCEL
        self.valence_condition = valence_condition.upper()
        self.all_affect_images = []

    def resolve_path(self, path_str):
        if not path_str:
            return None
        return os.path.join(self.project_root, path_str)

    def load_affect_images(self):
        if not os.path.exists(self.affect_excel):
            return []

        wb = load_workbook(self.affect_excel)
        sheet = wb.active
        affect_images = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            img_path, quadrant = row[1], row[3]
            if not img_path or quadrant is None:
                continue
            img_path = img_path.replace("\\", "/")
            full_path = self.resolve_path(img_path)
            if not full_path or not os.path.exists(full_path):
                continue
            if quadrant.startswith(self.valence_condition):
                affect_images.append({"path": full_path, "quadrant": quadrant})

        random.shuffle(affect_images)
        return affect_images

    def _match_fallback(self, path, ext):
        if not path:
            return None
        folder, fname = os.path.split(path)
        base = fname.split("_")[0]
        matches = glob.glob(os.path.join(folder, f"{base}*{ext}"))
        return matches[0] if matches else None

    def _fix_video(self, video_file):
        if not video_file:
            return None
        video_file = video_file.strip().replace("\\", "/")
        if os.path.exists(video_file):
            return video_file

        folder, fname = os.path.split(video_file)
        base = re.sub(r"[^a-zA-Z0-9\-]+", "_", os.path.splitext(fname)[0]) 
        candidates = glob.glob(os.path.join(folder, "*.mp4"))
        for c in candidates:
            c_base = os.path.splitext(os.path.basename(c))[0]
            if base in c_base or c_base in base:
                return c
        print(f"[WARN] No fallback match for {video_file}")
        return None

    def generate_dummy_trials(self,n=3):
        dummy_trials = []
        for i in range(n):
            dummy_trials.append({
                "video": None, 
                "spoof_segment_times": "1.0 - 2.0",
                "duration": 5.0,
                "affect_image": None,
                "quadrant": "LV-TEST",
                "trial_number": i+1
            })
        return dummy_trials

    def load_trials(self):
        trials = []
        if not os.path.exists(self.stimuli_excel):
            return self.generate_dummy_trials()

        wb = load_workbook(self.stimuli_excel)
        sheet = wb.active
 
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            try:
                video_file = row[8]
                spoof_times = row[5]
                duration = row[6]
                label = row[4]
            except Exception:
                continue

            video_path = self.resolve_path(video_file)
            video_path = self._fix_video(video_path)
            if not video_path:
                trials.append(self.generate_dummy_trials(1)[0])
                continue

            trial = {
                "video": video_path,
                "label": label,
                "spoof_segment_times": spoof_times or "",
                "duration": parse_duration(duration),
            }

            trial["trial_number"] = i + 1
            trials.append(trial)

        random.shuffle(trials)
        return trials